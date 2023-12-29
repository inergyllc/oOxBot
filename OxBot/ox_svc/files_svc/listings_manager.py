import json
import openpyxl
import re
import os
import datetime
import shutil
from openpyxl.utils import get_column_letter

# Global configuration to enable/disable features
global_config = {
    'process_branches': True,
    'process_people': True,
    'process_geocode': True,
    'process_listing': True
}

# Helper function to convert camelCase or PascalCase to snake_case
def to_snake_case(name):
    s1 = re.sub('(.)([A-Z][a-z]+)', r'\1_\2', name)
    return re.sub('([a-z0-9])([A-Z])', r'\1_\2', s1).lower()

# Generate output file path with timestamping

def generate_output_file_path(original_file_path, new_subfolder=None):
    base_folder = os.path.dirname(original_file_path)
    base_name = os.path.basename(original_file_path)
    name_part, ext_part = os.path.splitext(base_name)
    
    # Append the new_subfolder to the base_folder if provided
    if new_subfolder:
        base_folder = os.path.join(base_folder, new_subfolder)
    
    # Ensure the specified subdirectory exists
    if not os.path.exists(base_folder):
        os.makedirs(base_folder)
    
    # Generate timestamped file name
    timestamp = datetime.datetime.now().strftime("%Y%m%d%H%M%S")
    new_file_name = f"{name_part}_{timestamp}.xlsx"
    new_file_path = os.path.join(base_folder, new_file_name)
    
    return new_file_path

class GeocodeManager:
    def __init__(self, geocode_data, branch_counter,excluded_fields = None):
        self.geocode_data = geocode_data
        self.branch_counter = branch_counter
        self.raw_excluded_fields = excluded_fields
        self.excluded_fields = [ef['field'] for ef in excluded_fields if ef['type'] == 'geocode'] or []

    def flatten(self):
        # Process only the first result and specific fields
        result = self.geocode_data['results'][0] if self.geocode_data['results'] else {}
        address = result.get('address', {})
        position = result.get('position', {})

        # Flatten the geocode information
        flat_geocode = {
            f'branch_{self.branch_counter}_municipality': address.get('municipality', ''),
            f'branch_{self.branch_counter}_postal_code': address.get('postalCode', ''),
            f'branch_{self.branch_counter}_lat': position.get('lat', ''),
            f'branch_{self.branch_counter}_lon': position.get('lon', '')
        }
        return flat_geocode


class PeopleManager:
    def __init__(self, people_data, branch_counter, excluded_fields=None):
        self.people_data = people_data
        self.branch_counter = branch_counter
        self.raw_excluded_fields = excluded_fields
        self.excluded_fields = [ef['field'] for ef in excluded_fields if ef['type'] == 'people'] or []

    def flatten(self):
        flat_people = {}
        for counter, person in enumerate(self.people_data, start=1):
            for key, value in person.items():
                if key not in self.excluded_fields:  # Exclude 'branch' fields
                    flat_people[f'person_{self.branch_counter}.{counter}_{to_snake_case(key)}'] = value
        return flat_people


class BranchesManager:
    def __init__(self, branches_data, excluded_fields=None):
        self.branches_data = branches_data
        self.raw_self_excludes = excluded_fields
        self.excluded_fields = [ef['field'] for ef in excluded_fields if ef['type'] == 'branch'] or []

    def flatten_branch(self, branch_data, branch_counter):
        flat_branch = {}
        for key, value in branch_data.items():
            if key not in self.excluded_fields:  # Exclude 'branch' fields
                flat_branch[f'branch_{branch_counter}_{to_snake_case(key)}'] = value

        # Manage geocode and people separately
        if not global_config['process_geocode']:
            if 'geocode' in branch_data:
                geocode_manager = GeocodeManager(branch_data['geocode'], branch_counter, self.raw_self_excludes)
                flat_branch.update(geocode_manager.flatten())

        if not global_config['process_people']:
            if 'people' in branch_data:
                people_manager = PeopleManager(branch_data['people'], branch_counter, self.raw_self_excludes)
                flat_branch.update(people_manager.flatten())

        return flat_branch

    def flatten(self):
        all_branches = {}
        for counter, branch in enumerate(self.branches_data, start=1):
            all_branches.update(self.flatten_branch(branch, counter))
        return all_branches

class ListingManager:
    # Specific values as class variable, only created once for the class
    specific_values = [
        "Fueling the Future, Today",
         "Pioneering Energy Solutions for Tomorrow",
         "Driven by Innovation, Powered by Nature",
         "Where Energy Meets Excellence",
         "Sustaining the World, One Drop at a Time",
         "Empowering Progress with Every Barrel",
         "Beyond Oil - Crafting Energy Stories",
         "Commitment to Energy, Commitment to You",
         "Bringing Energy Closer to You",
         "Exploration Innovation Transformation",
         "Crafting the Curve of Energy Evolution",
         "The Power of Persistence in Every Drop",
         "From Deep Sea to Desert - We Deliver",
         "Bridging Resources, Building Tomorrow",
         "Nature's Power, Our Passion",
         "Redefining Energy Frontiers",
         "Your Global Energy Partner",
         "From Source to Solution - We're There",
         "Discovering Potential, Delivering Promise",
         "Change the Way You Think About Energy"
     ]

    def __init__(self, listing_data, excluded_fields=None):
        self.listing = listing_data
        self.raw_excluded_fields = excluded_fields
        self.excluded_fields = [ef['field'] for ef in (excluded_fields or []) if ef['type'] == 'listing']
        self.field_fns = [ 
            {
                'field': 'logo', 
                'type': 'contains', 
                'checkfor': 'generic', 
            },
            {
                'field': 'website_url', 
                'type': 'contains', 
                'checkfor': 'unknown-company-website.html', 
            }
        ]

    def process_contains_field(self, value, checkfor):
        if checkfor in value:
            return None
        else:
            return value
    
    def process_slogan(self, value):
        if value in self.specific_values:
            return None
        else:
            return value

    def process(self):
        flat_data = {}
        for key, value in self.listing.items():
            snake_key = to_snake_case(key)

            # Continue processing branches regardless of exclusion
            if snake_key == 'branches':
                if not global_config['process_branches']:
                    branches_manager = BranchesManager(value, self.raw_excluded_fields)
                    flat_data.update(branches_manager.flatten())
                continue

            # Exclude fields as needed
            if snake_key in self.excluded_fields:
                continue

            # Check if the field has a special processing rule or is 'slogan'
            if snake_key == 'slogan':
                value = self.process_slogan(value)
            else:
                for field_fn in self.field_fns:
                    if snake_key == field_fn['field'] and field_fn['type'] == 'contains':
                        value = self.process_contains_field(value, field_fn['checkfor'])

            # Handle array values with pipe symbol
            if isinstance(value, list):
                flat_data[snake_key] = '|'.join(map(str, value))
            else:
                flat_data[snake_key] = value

        return flat_data

class ListingsToExcel:
    def __init__(self, filename, excluded_fields=None, output_xlsx=None, max_rows_flattened=float('inf')):
        self.filename = filename
        self.max_rows_flattened = max_rows_flattened 
        self.excluded_fields = excluded_fields or []
        self.workbook = openpyxl.Workbook()
        self.sheet = self.workbook.active
        self.output_xlsx = output_xlsx or self.generate_default_output_name()

    def generate_default_output_name(self):
        base_folder, name = os.path.split(self.filename)
        name_part, ext_part = os.path.splitext(name)
        return os.path.join(base_folder, f"{name_part}.xlsx")
    
    # Check if the output file already exists and handle moving it to a backup location
    def save_xlsx(self):
        if os.path.exists(self.output_xlsx):
            backup_file_path = generate_output_file_path(self.output_xlsx)
            shutil.move(self.output_xlsx, backup_file_path)  # Move old xlsx to ts new
            print("Moved old file to: {}".format(backup_file_path))
            
        # Save the workbook to the original or specified path
        self.workbook.save(self.output_xlsx)
        print("Saved to: {}".format(self.output_xlsx))
    

    def process_listings(self):
        with open(self.filename, 'r') as file:
            data = json.load(file)
        listings = data['listings']

        row_num = 1
        for listing in listings:
            if row_num > self.max_rows_flattened:
                break
            col_num = 1
            if not global_config['process_listing']:
                listing_manager = ListingManager(listing, self.excluded_fields)
                flat_data = listing_manager.process()

            # Write data to the sheet
            for key, value in flat_data.items():
                self.sheet[f'{get_column_letter(col_num)}{row_num}'] = key
                self.sheet[f'{get_column_letter(col_num)}{row_num+1}'] = value
                col_num += 1

            row_num += 2  # Skip a row after each listing for better separation

        self.save_xlsx()

            
# Example usage
# listings_json_file_path = "E:/oxai/data/current/listings/listing_master_mirror.json"
# listings_excel_folder = "E:/oxai/data/current/listings"
# listings_excel_backup_folder = "E:/oxai/data/current/listings/backup"
# excluded_fields = ['owconnect_category_id']
# listings_to_excel = ListingsToExcel("E:/oxai/data/current/listings/listing_master_mirror.json", "assistant", excluded_fields=excluded_fields)
# listings_to_excel.process_listings()