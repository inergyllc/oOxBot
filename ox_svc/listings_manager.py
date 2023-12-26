import json
import openpyxl
import re
from openpyxl.utils import get_column_letter

# Assuming the file path as given


# Helper function to convert camelCase or PascalCase to snake_case
def to_snake_case(name):
    s1 = re.sub('(.)([A-Z][a-z]+)', r'\1_\2', name)
    return re.sub('([a-z0-9])([A-Z])', r'\1_\2', s1).lower()


class GeocodeManager:
    def __init__(self, geocode_data, branch_counter):
        self.geocode_data = geocode_data
        self.branch_counter = branch_counter

    def flatten(self):
        # Process only the first result and specific fields
        result = self.geocode_data['results'][0] if self.geocode_data['results'] else {}
        address = result.get('address', {})
        position = result.get('position', {})

        # Flatten the geocode information
        flat_geocode = {
            f'branch_{self.branch_counter}_municipality': address.get('municipality', ''),
            f'branch_{self.branch_counter}_postal_code': address.get('postalCode', ''),
            f'branch_{self.branch_counter}_lat': position.get('lat', ''),
            f'branch_{self.branch_counter}_lon': position.get('lon', '')
        }
        return flat_geocode


class PeopleManager:
    def __init__(self, people_data, branch_counter):
        self.people_data = people_data
        self.branch_counter = branch_counter

    def flatten(self):
        flat_people = {}
        for counter, person in enumerate(self.people_data, start=1):
            for key, value in person.items():
                flat_people[f'person_{self.branch_counter}.{counter}_{to_snake_case(key)}'] = value
        return flat_people


class BranchesManager:
    def __init__(self, branches_data):
        self.branches_data = branches_data

    def flatten_branch(self, branch_data, branch_counter):
        flat_branch = {}
        for key, value in branch_data.items():
            if key != "geocode" and key != "people":  # Exclude complex structures
                flat_branch[f'branch_{branch_counter}_{to_snake_case(key)}'] = value

        # Manage geocode and people separately
        if 'geocode' in branch_data:
            geocode_manager = GeocodeManager(branch_data['geocode'], branch_counter)
            flat_branch.update(geocode_manager.flatten())

        if 'people' in branch_data:
            people_manager = PeopleManager(branch_data['people'], branch_counter)
            flat_branch.update(people_manager.flatten())

        return flat_branch

    def flatten(self):
        all_branches = {}
        for counter, branch in enumerate(self.branches_data, start=1):
            all_branches.update(self.flatten_branch(branch, counter))
        return all_branches


class ListingManager:
    def __init__(self, listing_data, excluded_fields=None):
        self.listing = listing_data
        self.excluded_fields = excluded_fields or []

    def process(self):
        flat_data = {}
        for key, value in self.listing.items():
            key = to_snake_case(key)

            if key in self.excluded_fields:
                continue

            if key == 'branches':
                branches_manager = BranchesManager(value)
                flat_data.update(branches_manager.flatten())
            elif isinstance(value, list):
                flat_data[key] = '|'.join(value)
            else:
                flat_data[key] = value
        return flat_data

class ListingsToExcel:
    def __init__(self, filename, excluded_fields=None, max_rows_flattened = float('inf')):
        self.filename = filename
        self.max_rows_flattened = max_rows_flattened 
        self.excluded_fields = excluded_fields or []
        self.workbook = openpyxl.Workbook()
        self.sheet = self.workbook.active

    def process_listings(self):
        with open(self.filename, 'r') as file:
            data = json.load(file)
        listings = data['listings']

        row_num = 1
        for listing in listings:
            if row_num > self.max_rows_flattened:
                break
            col_num = 1
            listing_manager = ListingManager(listing, self.excluded_fields)
            flat_data = listing_manager.process()

            # Write data to the sheet
            for key, value in flat_data.items():
                self.sheet[f'{get_column_letter(col_num)}{row_num}'] = key
                self.sheet[f'{get_column_letter(col_num)}{row_num+1}'] = value
                col_num += 1

            row_num += 1  # Can skip a row after each listing for better separation

        # Save the workbook
        self.workbook.save("output.xlsx")

# Example usage
# listings_json_file_path = "E:/oxai/data/current/listings/listing_master_mirror.json"
# excluded_fields = ['owconnect_category_id']
# listings_to_excel = ListingsToExcel("E:/oxai/data/current/listings/listing_master_mirror.json", excluded_fields=excluded_fields)
# listings_to_excel.process_listings()


